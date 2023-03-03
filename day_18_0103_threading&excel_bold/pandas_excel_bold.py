import pandas as pd
from openpyxl import Workbook  # https://openpyxl.readthedocs.io/en/stable/index.html
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows  # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.dataframe.html


# Load data from the input Excel file
df = pd.read_excel('input_file.xlsx')

# Calculate the sum of columns 1 and 2 and append it to column 3
df['SUM_OF'] = df['Column 0'] + df['Column 1']

# Save the updated data to the output Excel file
df.to_excel('input_file_tmp.xlsx', index=False)


# Load data from the input Excel file
df_new = pd.read_excel('input_file_tmp.xlsx')


# Specify which columns to make bold
bold_cols = ['Column 0', 'Column 2', 'SUM_OF']

# Create a new Excel workbook to write the updated data to
wb = Workbook()
ws = wb.active

# Write the data to the worksheet
for r in dataframe_to_rows(df_new, index=False, header=True):
    ws.append(r)

# Apply bold formatting to the specified columns
for col in bold_cols:
    col_num = df.columns.get_loc(col) #jak wczytujemy z xlsx nie ma indeksu + 1
    for cell in ws[1:ws.max_row]:  # Exclude header row
        cell[col_num].font = Font(bold=True)

# Save the updated data to the output Excel file
wb.save('output_file.xlsx')
